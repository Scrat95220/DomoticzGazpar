#!/usr/bin/python3
# -*- coding: utf-8 -*-
# (C) v1.4.1 2022-01-21 Scrat
"""Generates energy consumption JSON files from GRDf consumption data
collected via their  website (API).
"""

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

import openpyxl
import urllib.request
import ssl
import base64
import configparser
import os
import requests
import datetime
import logging
import json
from dateutil.relativedelta import relativedelta
from urllib.parse import urlencode
import http.cookiejar

LOGIN_BASE_URI = 'https://login.monespace.grdf.fr/sofit-account-api/api/v1/auth'
API_BASE_URI = 'https://monespace.grdf.fr/'

logLevel = "INFO"

base64string=""

userName = ""
password = ""
devicerowid = ""
devicerowidm3 = ""
nbDaysImported = 30
dbPath = ""
xlsPath = ""
domoticzserver   = ""
domoticzusername = ""
domoticzpassword = ""

script_dir=os.path.dirname(os.path.realpath(__file__)) + os.path.sep

class GazparServiceException(Exception):
    """Thrown when the webservice threw an exception."""
    pass

def domoticzrequest (url):
  global base64string

  context = ssl._create_unverified_context()
  request = urllib.request.Request(domoticzserver + url)
  if(domoticzusername != "" and domoticzpassword!= ""):
    base64string = base64.encodebytes(('%s:%s' % (domoticzusername, domoticzpassword)).encode()).decode().replace('\n', '')
    request.add_header("Authorization", "Basic %s" % base64string)
    
  try:
    logging.debug("Domoticz Request : \n" + request.full_url)  
    response = urllib.request.urlopen(request, context=context)
    logging.debug("Domoticz Response : \n" + response.read().decode('utf-8')) 
    return response.read()
  except urllib.error.HTTPError as e:
    print(e.__dict__)
    logging.error("Domoticz call - HttpError :"+str(e.__dict__)+'\n')
    exit()
  except urllib.error.URLError as e:
    print(e.__dict__)
    logging.error("Domoticz call - UrlError :"+str(e.__dict__)+'\n')
    exit()
  
# Date formatting 
def dtostr(date):
    return date.strftime("%Y-%m-%d")
    
def login():
    """Logs the user into the GRDF API.
    """
    session = requests.Session()
    SESSION_TOKEN_URL = "https://connexion.grdf.fr/api/v1/authn"
    SESSION_TOKEN_PAYLOAD = """{{
                "username": "{0}",
                "password": "{1}",
                "options": {{
                    "multiOptionalFactorEnroll": "false",
                    "warnBeforePasswordExpired": "false"
                }}
            }}"""
    AUTH_TOKEN_URL = "https://connexion.grdf.fr/login/sessionCookieRedirect"
    AUTH_TOKEN_PARAMS = """{{
                "checkAccountSetupComplete": "true",
                "token": "{0}",
                "redirectUrl": "https://monespace.grdf.fr"
            }}"""
    # self.session = Session()
    session.headers.update({"domain": "grdf.fr"})
    session.headers.update({"Content-Type": "application/json"})
    session.headers.update({"X-Requested-With": "XMLHttpRequest"})

    payload = SESSION_TOKEN_PAYLOAD.format(userName, password)

    response = session.post(SESSION_TOKEN_URL, data=payload)

    if response.status_code != 200:
        raise Exception(
            f"An error occurred while logging in. Status code: {response.status_code} - {response.text} - {response.payload}")
        exit()

    # get auth token
    session_token = response.json().get("sessionToken")
    logging.debug("Session token: %s", session_token)
    jar = http.cookiejar.CookieJar()
    # self.session = Session()
    session.headers.update({"Content-Type": "application/json"})
    session.headers.update({"X-Requested-With": "XMLHttpRequest"})

    params = json.loads(AUTH_TOKEN_PARAMS.format(session_token))

    response = session.get(AUTH_TOKEN_URL, params=params, allow_redirects=True, cookies=jar)

    if response.status_code != 200:
        raise Exception(
            f"An error occurred while getting the auth token. Status code: {response.status_code} - {response.text}")
        exit()

    auth_token = session.cookies.get("auth_token", domain="monespace.grdf.fr")

    # Create a session.
    # self.session = Session()
    session.headers.update({"Host": "monespace.grdf.fr"})
    session.headers.update({"Domain": "grdf.fr"})
    session.headers.update({"X-Requested-With": "XMLHttpRequest"})
    session.headers.update({"Accept": "application/json"})
    session.cookies.set("auth_token", auth_token, domain="monespace.grdf.fr")

    return session
    
def update_counters(session, start_date, end_date):
    logging.debug('start_date: ' + start_date + "; end_date: " + end_date)
    
    #3nd request- Get NumPCE
    resp3 = session.get('https://monespace.grdf.fr/api/e-connexion/users/pce/historique-consultation')
    logging.debug("Get NumPce Response : \n" + resp3.text)
    if resp3.status_code != requests.codes.ok:
        print("Get NumPce call - error status :",resp3.status_code, '\n');
        logging.error("Get NumPce call - error status :",resp3.status_code, '\n')
        exit()
    
    j = json.loads(resp3.text)
    numPce = j[0]['numPce']
    
    data = get_data_with_interval(session, 'Mois', numPce, start_date, end_date)
    
    j = json.loads(data)
    #print (j)
    index = j[str(numPce)]['releves'][0]['indexDebut']      
    #print(index)
    
    for releve in j[str(numPce)]['releves']:
        #print(releve)
        req_date = releve['journeeGaziere']
        conso = releve['energieConsomme']
        indexm3 = releve['indexFin']
        coeffConversion = releve['coeffConversion']
        req_date_time = releve['dateFinReleve']
        volume = releve['volumeBrutConsomme']
        
        try :
            #volume = round(int(conso)/int(coeffConversion),2)
            index = index + conso
        except TypeError:
            print(req_date, conso, index, "Invalid Entry")
            continue;
        
        date_time = datetime.datetime.strptime(req_date_time[:19], '%Y-%m-%dT%H:%M:%S').strftime("%Y-%m-%d %H:%M:%S")
        
        #print(req_date, conso, index)
        if devicerowid:
            logging.debug("Data to inject : " + req_date + ";" + devicerowid + ";" + str(int(conso)*1000) + ";" + str(index))
            
            # Generate URLs, for historique and for update
            args = {'type': 'command', 'param': 'udevice', 'idx': devicerowid, 'svalue': str(index) + ";" + str(int(conso)*1000) + ";" + req_date}
            url_historique = '/json.htm?' + urlencode(args)
             
            args['svalue'] = str(index)  + ";" + str(int(conso)*1000) + ";" + date_time
            url_daily = '/json.htm?' + urlencode(args)

            args['svalue'] = str(int(conso)*1000)
            url_current = '/json.htm?' + urlencode(args)
            
            domoticzrequest(url_historique)
            
        if devicerowidm3:
            logging.debug("Data to inject : " + req_date + ";" + devicerowidm3 + ";" + str(volume) + ";" + str(indexm3))
            
            # Generate URLs, for historique and for update
            args_m3 = {'type': 'command', 'param': 'udevice', 'idx': devicerowidm3, 'svalue': str(indexm3) + ";" + str(volume) + ";" + req_date}
            url_historique_m3 = '/json.htm?' + urlencode(args_m3)
            
            args_m3['svalue'] = str(indexm3)  + ";" + str(volume) + ";" + date_time
            url_daily_m3 = '/json.htm?' + urlencode(args_m3)

            args_m3['svalue'] = str(volume)
            url_current_m3 = '/json.htm?' + urlencode(args_m3)
            
            domoticzrequest(url_historique_m3)
    
    if devicerowid:
        domoticzrequest(url_current)
        domoticzrequest(url_daily)
    
    if devicerowidm3:
        domoticzrequest(url_current_m3)
        domoticzrequest(url_daily_m3)
        
def get_data_with_interval(session, resource_id, numPce, start_date=None, end_date=None):
    r=session.get('https://monespace.grdf.fr/api/e-conso/pce/consommation/informatives?dateDebut='+ start_date + '&dateFin=' + end_date + '&pceList[]=' + str(numPce))
    logging.debug("Data : \n" + r.text)
    if r.status_code != requests.codes.ok:
        print("Get data - error status :"+r.status_code+'\n');
        logging.error("Get data - error status :",r.status_code, '\n')
        exit()
    return r.text
    
def get_config():
    configuration_file = script_dir + '/domoticz_gazpar.cfg'
    config = configparser.ConfigParser(allow_no_value=True)
    config.read(configuration_file)
    global logLevel
    global userName
    global password
    global devicerowid
    global devicerowidm3
    global nbDaysImported
    global xlsPath
    global dbPath
    global domoticzserver
    global domoticzusername
    global domoticzpassword
    
    logLevel = config['SETTINGS']['LOG_LEVEL']
    userName = config['GRDF']['GAZPAR_USERNAME']
    password = config['GRDF']['GAZPAR_PASSWORD']
    devicerowid = config['DOMOTICZ']['DOMOTICZ_ID']
    devicerowidm3 = config['DOMOTICZ']['DOMOTICZ_ID_M3']
    nbDaysImported = config['GRDF']['NB_DAYS_IMPORTED']
    xlsPath = config['GRDF']['XLS_PATH']
    dbPath = config['DOMOTICZ_SETTINGS']['DB_PATH']
    domoticzserver   = config['DOMOTICZ_SETTINGS']['HOSTNAME']
    domoticzusername = config['DOMOTICZ_SETTINGS']['USERNAME']
    domoticzpassword = config['DOMOTICZ_SETTINGS']['PASSWORD']
    
    logging.debug("config : " + userName + "," + password + "," + devicerowid + "," + devicerowidm3 + "," + nbDaysImported + "," + xlsPath + "," + dbPath + "," + domoticzserver + "," + domoticzusername + "," + domoticzpassword)
    
def xlsimport():
    workbook = openpyxl.load_workbook(xlsPath)
    sheet = workbook.active 
    nbRows = sheet.max_row
    index = sheet.cell(row=10, column=3).value
    
    if(int(nbDaysImported)>nbRows):
        print("nbDaysImported max allowed is " + str(nbRows-9))
        logging.debug("nbDaysImported max allowed is " + str(nbRows-9))
        exit()
        
    #i : inital date to import
    i=nbRows-int(nbDaysImported)+1
    while nbRows+1 > i:
        req_date = sheet.cell(row=i, column=2).value
        conso = sheet.cell(row=i, column=6).value
        indexm3 = sheet.cell(row=i, column=4).value
        req_date_time = sheet.cell(row=i, column=2).value 
        volume = sheet.cell(row=i, column=5).value 
        
        i = i+1
        
        try :
            index = index + conso
        except TypeError:
            print(req_date, conso, index, "Invalid Entry")
            continue;
            
        date_time = datetime.datetime.strptime(req_date_time, '%d/%m/%Y').strftime("%Y-%m-%d %H:%M:%S")
        req_date = datetime.datetime.strptime(req_date, '%d/%m/%Y').strftime("%Y-%m-%d")
        
        #print(req_date, conso, index)
        if devicerowid:
            logging.debug("Data to inject : " + req_date + ";" + devicerowid + ";" + str(int(conso)*1000) + ";" + str(index))
            
            # Generate URLs, for historique and for update
            args = {'type': 'command', 'param': 'udevice', 'idx': devicerowid, 'svalue': str(index) + ";" + str(int(conso)*1000) + ";" + req_date}
            url_historique = '/json.htm?' + urlencode(args)
             
            args['svalue'] = str(index)  + ";" + str(int(conso)*1000) + ";" + date_time
            url_daily = '/json.htm?' + urlencode(args)

            args['svalue'] = str(int(conso)*1000)
            url_current = '/json.htm?' + urlencode(args)
            
            #print(url_historique)
            domoticzrequest(url_historique)
            
        if devicerowidm3:
            logging.debug("Data to inject : " + req_date + ";" + devicerowidm3 + ";" + str(volume) + ";" + str(indexm3))
            
            # Generate URLs, for historique and for update
            args_m3 = {'type': 'command', 'param': 'udevice', 'idx': devicerowidm3, 'svalue': str(indexm3) + ";" + str(volume) + ";" + req_date}
            url_historique_m3 = '/json.htm?' + urlencode(args_m3)
            
            args_m3['svalue'] = str(indexm3)  + ";" + str(volume) + ";" + date_time
            url_daily_m3 = '/json.htm?' + urlencode(args_m3)

            args_m3['svalue'] = str(volume)
            url_current_m3 = '/json.htm?' + urlencode(args_m3)
            
            domoticzrequest(url_historique_m3)
            
    
    if devicerowid:
        domoticzrequest(url_current)
        domoticzrequest(url_daily)
    
    if devicerowidm3:
        domoticzrequest(url_current_m3)
        domoticzrequest(url_daily_m3)
        

# Main script 
def main():
    logging.basicConfig(filename=script_dir + '/domoticz_gazpar.log', format='%(asctime)s %(message)s', filemode='w') 

    try:
        # Get Configuration
        logging.info("Get configuration")
        get_config()
        
        #Set log level
        if(logLevel == "INFO"):
            logging.getLogger().setLevel(logging.INFO)
        if(logLevel == "DEBUG"):
            logging.getLogger().setLevel(logging.DEBUG)
        if(logLevel == "ERROR"):
            logging.getLogger().setLevel(logging.ERROR)            
        
        # Login to GRDF API
        logging.info("logging in as %s...", userName)
        token = login()
        logging.info("logged in successfully!")

        today = datetime.date.today()

        # Update Counters Domoticz
        logging.info("retrieving data...")
        update_counters(token, dtostr(today - relativedelta(days=int(nbDaysImported))), \
                                             dtostr(today))
                                             
        logging.info("got data!")
    except GazparServiceException as exc:
        logging.error(exc)
        exit()

if __name__ == "__main__":
    main()
